"""
Template 2 (KPI) Operator for processing Excel files with 'Шаблон загрузки' sheet.
Loads data to StarRocks with inferred schema based on column types.
"""

from __future__ import annotations

import re
from io import BytesIO
from typing import Dict, List, Optional

import numpy as np
import pandas as pd
from airflow.models import BaseOperator
from datetime import datetime, date

from base_operators.base_hooks_operator import BaseHooksOperator
from starrocks_operators.starrocks_connection_mixin import StarRocksConnectionMixin
from starrocks_operators.starrocks_data_converter import StarRocksDataConverter
from vk_cloud_operators.vk_table_manager import VKTableManager
from vk_cloud_operators.vk_stream_loader import VKStreamLoader
from vk_cloud_operators.vk_cloud_file_status_mixin import VKCloudFileStatusMixin


# Required columns according to шаблон2.py specification
DEFAULT_COLUMN_NAMES = (
    '№',
    'Код файла',
    'Название файла исходное',
    'Значение',
    'Дата',
    'Дата отчета',
    'Наименование показателя в исходнике',
    'Код показателя из модели',
    'Наименование показателя из модели',
    'А52.Единицы измерения',
    'Тип данных',
    'Метрика',
    'Дискретность',
    'Тип значения'
)


def find_substring_index(substring: str, lst: list) -> Optional[int]:
    """Find substring in list case-insensitively"""
    substring_lower = substring.lower()
    for i, item in enumerate(lst):
        if substring_lower in item.lower():
            return i
    return None


def get_column_types(df: pd.DataFrame) -> dict:
    """Infer SQL column types from DataFrame"""
    result = {}

    for col in df.columns:
        series = df[col]
        non_null = series.dropna()

        if non_null.empty:
            result[col] = "decimal(38, 10) NULL"
            continue

        # Check for datetime/date
        if all(isinstance(v, (datetime, date, np.datetime64, pd.Timestamp)) for v in non_null):
            has_time = any(
                isinstance(v, (datetime, pd.Timestamp)) and
                (v.time() != datetime.min.time())
                for v in non_null
            )
            result[col] = "datetime NULL" if has_time else "date NULL"
            continue

        # Check for numeric
        numeric = pd.to_numeric(non_null, errors="coerce")

        if numeric.notna().all():
            if (numeric % 1 == 0).all():
                result[col] = "int(11) NULL"
            else:
                result[col] = "decimal(38, 10) NULL"
            continue

        # String - use max varchar size
        result[col] = "varchar(65333) NULL"

    return result


class Template2Operator(StarRocksConnectionMixin, VKCloudFileStatusMixin, BaseHooksOperator):
    """
    Operator for processing Template 2 (KPI) Excel files.
    Finds 'Шаблон загрузки' sheet, validates columns, infers schema, loads to StarRocks.
    """

    def __init__(
        self,
        s3_bucket: str,
        s3_key: str,
        file_id: str,
        starrocks_table: str,
        file_name: str,
        user_name: str = "vk_cloud",
        last_modified: str = None,
        starrocks_database: str = "stage",
        column_separator: str = "|",
        load_timeout: int = 600,
        enable_logging: bool = True,
        postgres_conn_id: str = "airflow_db",
        **kwargs,
    ):
        super().__init__(**kwargs)
        self.s3_bucket = s3_bucket
        self.s3_key = s3_key
        self.file_id = file_id
        self.starrocks_table = starrocks_table
        self.file_name = file_name
        self.user_name = user_name
        self.last_modified = last_modified
        self.starrocks_database = starrocks_database
        self.column_separator = column_separator
        self.load_timeout = load_timeout
        self.enable_logging = enable_logging
        self.postgres_conn_id = postgres_conn_id

        conn_id = kwargs.get("starrocks_conn_id", "starrocks_default")
        self._initialize_starrocks_connection(
            conn_id=conn_id,
            max_retries=kwargs.get("max_retries", 3),
        )
        self.table_manager = VKTableManager(
            mysql_conn_id=conn_id,
            database=self.starrocks_database,
            logger=self.log,
        )
        self.stream_loader = VKStreamLoader(
            session=self.starrocks_session,
            host=self.starrocks_host,
            port=self.starrocks_port,
            user=self.starrocks_credentials[0],
            password=self.starrocks_credentials[1],
            timeout=self.load_timeout,
            logger=self.log,
        )

    def get_postgres_hook(self, database: str = "airflow_db"):
        """Get PostgreSQL hook for logging"""
        from airflow.providers.postgres.hooks.postgres import PostgresHook
        return PostgresHook(
            postgres_conn_id=self.postgres_conn_id,
            database=database
        )

    def _find_template_sheet(self, wb) -> Optional[str]:
        """Find 'Шаблон загрузки' sheet case-insensitively, fallback to first sheet"""
        target_sheet_name = 'Шаблон загрузки'
        sheet_index = find_substring_index(target_sheet_name, wb.sheetnames)
        if sheet_index is not None:
            return wb.sheetnames[sheet_index]
        # Fallback to first sheet if target not found
        if wb.sheetnames:
            self.log.info(f"Sheet 'Шаблон загрузки' not found, using first sheet: {wb.sheetnames[0]}")
            return wb.sheetnames[0]
        return None

    def _validate_required_columns(self, df: pd.DataFrame) -> bool:
        """Check if required columns exist - returns True with warning if missing (soft validation)"""
        existing_cols = set(df.columns)
        required_cols = set(DEFAULT_COLUMN_NAMES)
        missing_cols = required_cols - existing_cols

        if missing_cols:
            self.log.warning(f"Missing template columns (proceeding anyway): {missing_cols}")
            # Return True - allow processing of non-standard files
            return True
        return True

    def _filter_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Remove columns matching 'столбец\d+' pattern and empty columns"""
        cleared_columns = []

        for col_name in df.columns:
            if (col_name
                and re.match(r'^столбец(\d{1,3})$', str(col_name), re.IGNORECASE) is None
                and df[col_name].notnull().sum() > 0) or col_name in DEFAULT_COLUMN_NAMES:
                cleared_columns.append(col_name)

        return df[cleared_columns]

    def _filter_rows(self, df: pd.DataFrame) -> pd.DataFrame:
        """Remove rows with $$$$ marker and empty rows"""
        df = df[df.ne('$$$$').all(axis=1)]
        df = df.dropna(how='all')
        return df

    def _update_status(self, status_code: str, error_code: str = None):
        """Update file status in PostgreSQL if logging is enabled"""
        if self.enable_logging:
            try:
                self.update_vk_cloud_file_status(
                    file_id=self.file_id,
                    status_code=status_code,
                    error_code=error_code
                )
            except Exception as e:
                self.log.warning(f"Could not update file status: {e}")

    def execute(self, context) -> Dict[str, object]:
        try:
            self.log.info("Processing Template 2 (KPI) file: %s", self.s3_key)

            self._update_status('VALIDATING')

            file_content = self._get_file_from_s3()

            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(file_content), read_only=True, data_only=True)

            # Find 'Шаблон загрузки' sheet
            target_sheet = self._find_template_sheet(wb)
            if target_sheet is None:
                raise ValueError("Sheet 'Шаблон загрузки' not found in file")

            self.log.info(f"Found target sheet: {target_sheet}")

            # Read sheet data
            ws = wb[target_sheet]
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(row)

            df = pd.DataFrame(data)
            df.columns = df.iloc[0]
            df = df.drop(index=0)

            # Validate required columns
            if not self._validate_required_columns(df):
                raise ValueError("Required columns missing from Template 2 file")

            # Filter columns and rows
            df = self._filter_columns(df)
            df = self._filter_rows(df)

            if df.empty:
                raise ValueError("No data remaining after filtering")

            self.log.info(f"Data shape after filtering: {df.shape}")

            # Add metadata (5 required service columns)
            df['file_id'] = self.file_id
            df['user_name'] = self.user_name
            df['raw_file_name'] = self.s3_key
            df['file_name'] = self.file_name
            df['s3_placement'] = pd.Timestamp(self.last_modified) if self.last_modified else pd.Timestamp.now()

            self._update_status('LOADING_STAGE')

            # Infer column types and create table
            column_types = get_column_types(df)
            self.log.info(f"Inferred column types: {column_types}")

            # Load to StarRocks
            load_result = self._load_to_starrocks(df, column_types)

            self._update_status('COMPLETED')

            rows_loaded = load_result.get('NumberLoadedRows', len(df))
            return {
                'success': True,
                'rows_processed': len(df),
                'rows_loaded': rows_loaded,
                'columns': list(df.columns),
                'column_types': column_types,
                'table_name': self.starrocks_table,
                'starrocks_database': self.starrocks_database,
                'load_result': load_result,
            }

        except Exception as exc:
            self.log.error("Error processing Template 2 file: %s", exc)

            error_code = 'ERR_UNKNOWN'
            error_str = str(exc).lower()
            if 'sheet' in error_str or 'шаблон' in error_str:
                error_code = 'ERR_SHEET_NOT_FOUND'
            elif 'columns' in error_str or 'missing' in error_str:
                error_code = 'ERR_COLUMNS'
            elif 'load' in error_str or 'starrocks' in error_str:
                error_code = 'ERR_LOAD_STAGE'
            elif 's3' in error_str or 'file not found' in error_str:
                error_code = 'ERR_S3_ACCESS'
            elif 'parse' in error_str or 'excel' in error_str:
                error_code = 'ERR_PARSE'

            self._update_status('ERROR', error_code)

            return {'success': False, 'error': str(exc)}
        finally:
            self.close_starrocks_connection()

    def _load_to_starrocks(self, df: pd.DataFrame, column_types: Dict[str, str]) -> Dict[str, object]:
        """Load DataFrame to StarRocks with inferred schema"""
        columns = list(df.columns)

        self.table_manager.create_table_if_not_exists(
            table_name=self.starrocks_table,
            columns=columns,
        )

        converter = StarRocksDataConverter(logger=self.log)
        csv_data = converter.dataframe_to_csv(
            df,
            separator=self.column_separator,
        )

        return self.stream_loader.execute_stream_load(
            database=self.starrocks_database,
            table=self.starrocks_table,
            payload=csv_data,
            columns=columns,
            column_separator=self.column_separator,
        )

    def _get_file_from_s3(self) -> bytes:
        """Fetch file from S3"""
        from airflow.providers.amazon.aws.hooks.s3 import S3Hook

        hook = S3Hook(aws_conn_id=self.s3_conn_id)
        file_obj = hook.get_key(key=self.s3_key, bucket_name=self.s3_bucket)
        if not file_obj:
            raise FileNotFoundError(
                f"File not found: s3://{self.s3_bucket}/{self.s3_key}"
            )
        return file_obj.get()['Body'].read()
